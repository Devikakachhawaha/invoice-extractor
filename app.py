import streamlit as st
import pdfplumber
import pandas as pd
import re
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="PDF Extractor", layout="wide")
st.title("📄 PDF Extractor")
st.write("Upload PDF files and extract invoice data into Excel.")


def extract_pdf_text(uploaded_file):
    full_text = ""
    page_texts = []
    try:
        with pdfplumber.open(uploaded_file) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
                    page_texts.append(text)
                else:
                    page_texts.append("")
    except Exception as e:
        st.error(f"PDF Error: {e}")
    return full_text, page_texts


def extract_data(text, page_texts):
    data = {
        "Invoice No. (from Shipping Bill)": None,
        "Port Code (from Shipping Bill)": None,
        "Shipping Bill No. (from Shipping Bill)": None,
        "Shipping Bill Date (from Shipping Bill)": None,
        "Invoice Date (from Shipping Bill)": None,
        "IGST VALUE": None,
        "IGST TAX AMOUNT": None,
        "FOB": None,
        "FREIGHT": None,
        "INSURANCE": None,
        "TAXABLE VALUE (FOB+FREIGHT+INSURANCE)": None,
        "LUT": None,
        "F_INV_NO": None,
        "F_INV_AMT": None,
        "F_CURRENCY": None,
        "EXCHANGE_RATE": None,
    }

    clean_text = re.sub(r'\s+', ' ', text)

    # --- Port Code, SB No, SB Date ---
    # Pattern: "INDIAN CUSTOMS EDI SYSTEM" header area contains Port Code / SB No / SB Date
    sb_match = re.search(
        r'INDIAN CUSTOMS EDI SYSTEM\s+([A-Z0-9]+)\s+(\d+)\s+([0-9A-Z\-]+)',
        clean_text
    )
    if sb_match:
        data["Port Code (from Shipping Bill)"] = sb_match.group(1)
        data["Shipping Bill No. (from Shipping Bill)"] = sb_match.group(2)
        data["Shipping Bill Date (from Shipping Bill)"] = sb_match.group(3)

    # ==========================================
    # STEP 1: Extract invoice number ONCE, from F.INVOICE SUMMARY (page 1).
    # This is the single source of truth for invoice number — reused below for
    # both the main table and the summary table, so they can never disagree.
    # Anchor: "1 <token> <amount> <currency>" right before the "LET EXPORT COPY"
    # watermark text (rendered reversed as "X E T E L..." due to pdfplumber's
    # column-jumbling on this rotated-label page layout).
    # ==========================================
    page1_text = re.sub(r'\s+', ' ', page_texts[0]) if page_texts else clean_text

    invoice_no = None
    f_inv_match = re.search(
        r'\b1\s+(\S+)\s+([\d,]+(?:\.\d+)?)\s+([A-Z]{3})\s+[XE]',
        page1_text
    )
    if not f_inv_match:
        # Fallback: same triplet without requiring the watermark anchor right
        # after — some PDFs may jumble slightly differently around that point.
        f_inv_match = re.search(
            r'\b1\s+(\S+)\s+([\d,]+(?:\.\d+)?)\s+([A-Z]{3})\b',
            page1_text
        )
    if f_inv_match:
        invoice_no = f_inv_match.group(1)
        data["F_INV_NO"] = invoice_no
        raw_amt = f_inv_match.group(2).replace(",", "")
        data["F_INV_AMT"] = round(float(raw_amt), 2)
        data["F_CURRENCY"] = f_inv_match.group(3).upper()

    # --- Invoice Date from PART-II A.REF ---
    # Now look up the date using the ALREADY-KNOWN invoice number (escaped for
    # regex) rather than an independent "1 <token> <date>" guess that could
    # match the wrong occurrence elsewhere in a multi-page document.
    if invoice_no:
        data["Invoice No. (from Shipping Bill)"] = invoice_no
        date_match = re.search(
            re.escape(invoice_no) + r'\s+(\d{2}/\d{2}/\d{4})',
            clean_text
        )
        if date_match:
            original_date = date_match.group(1)
            try:
                converted_date = datetime.strptime(original_date, "%d/%m/%Y").strftime("%d-%b-%y").upper()
                data["Invoice Date (from Shipping Bill)"] = converted_date
            except ValueError:
                data["Invoice Date (from Shipping Bill)"] = original_date
    else:
        # Fallback: F.INVOICE SUMMARY anchor failed (e.g. unexpected layout).
        # Try the original positional guess as a last resort.
        inv_match = re.search(r'\b1\s+(\S+)\s+(\d{2}/\d{2}/\d{4})', clean_text)
        if inv_match:
            data["Invoice No. (from Shipping Bill)"] = inv_match.group(1)
            data["F_INV_NO"] = inv_match.group(1)
            original_date = inv_match.group(2)
            try:
                converted_date = datetime.strptime(original_date, "%d/%m/%Y").strftime("%d-%b-%y").upper()
                data["Invoice Date (from Shipping Bill)"] = converted_date
            except ValueError:
                data["Invoice Date (from Shipping Bill)"] = original_date

    # --- FOB Value (LM pattern = assessed FOB in INR) ---
    # NOTE: pdfplumber extracts the rotated "C.VALU SUMMA" sidebar label as "LM"
    # immediately followed by the FOB figure. This pattern is correct — keep it.
    fob_match = re.search(r'LM\s+([0-9.]+)', clean_text)
    if fob_match:
        fob_value = round(float(fob_match.group(1)), 2)
        data["FOB"] = fob_value

    # --- Freight ---
    # The row "5.COM 2. IGST AMT <freight> <insurance> <discount> <commission>"
    # The FIRST number after "IGST AMT" is FREIGHT, the SECOND is INSURANCE.
    freight_match = re.search(
        r'5\.COM 2\. IGST AMT\s+([0-9]+(?:\.[0-9]+)?)',
        clean_text
    )
    if freight_match:
        v = float(freight_match.group(1))
        if v > 0:
            data["FREIGHT"] = round(v, 2)

    # --- Insurance ---
    insurance_match = re.search(
        r'5\.COM 2\. IGST AMT\s+[0-9]+(?:\.[0-9]+)?\s+([0-9]+(?:\.[0-9]+)?)',
        clean_text
    )
    if insurance_match:
        v = float(insurance_match.group(1))
        if v > 0:
            data["INSURANCE"] = round(v, 2)

    data["TAXABLE VALUE (FOB+FREIGHT+INSURANCE)"] = None

    # --- LUT detection: if 4.IGST VALUE is followed by a number -> LUT=N, else LUT=Y ---
    igst_val_present = re.search(r'4\.IGST VALUE\s+(\d[\d.]*)', clean_text)
    is_lut = not bool(igst_val_present)
    data["LUT"] = "Y" if is_lut else "N"

    # --- IGST VALUE: from "D. EX. PR." table, field "4.IGST VALUE" ---
    # Only present (non-zero) for non-LUT shipments. LUT shipments show no
    # number after this label, so igst_val_present will be None and the field
    # stays blank — consistent with the LUT detection above.
    if igst_val_present:
        data["IGST VALUE"] = round(float(igst_val_present.group(1)), 2)

    # --- IGST Amount: only for non-LUT shipments ---
    if not is_lut:
        igst_match = re.search(r'3\.CESS AMT\s+([0-9.]+)\s+([0-9.]+)', clean_text)
        if igst_match:
            igst_amt = round(float(igst_match.group(2)), 2)
            data["IGST TAX AMOUNT"] = igst_amt if igst_amt > 0 else None

    # ==========================================
    # EXCHANGE RATE — from page 2 text
    # ==========================================
    ex_rate = None
    if len(page_texts) >= 2:
        page2_text = re.sub(r'\s+', ' ', page_texts[1])

        ex_match = re.search(
            r'9\.EXCHANGE\s+RATE.*?1\s+[A-Z]{3}\s+INR\s+([\d,]+(?:\.\d+)?)',
            page2_text,
            re.IGNORECASE | re.DOTALL
        )
        if ex_match:
            ex_rate = round(float(ex_match.group(1).replace(",", "")), 4)

        if ex_rate is None:
            ex_match2 = re.search(
                r'1\s+[A-Z]{3}\s+INR\s+([\d,]+(?:\.\d+)?)',
                page2_text,
                re.IGNORECASE
            )
            if ex_match2:
                ex_rate = round(float(ex_match2.group(1).replace(",", "")), 4)

        if ex_rate is None:
            ex_match3 = re.search(
                r'(?:EXCHANGE\s+RATE)[^0-9]{0,80}INR\s+([\d,]+(?:\.\d+)?)',
                page2_text,
                re.IGNORECASE
            )
            if ex_match3:
                ex_rate = round(float(ex_match3.group(1).replace(",", "")), 4)

    data["EXCHANGE_RATE"] = ex_rate
    return data


def build_excel(df, summary_rows):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_cell(cell, bold=True, font_color="000000", fill_color=None,
                   wrap=True, size=9, italic=False, h_align="center"):
        cell.font = Font(bold=bold, color=font_color, size=size, italic=italic)
        cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
        cell.border = bdr
        if fill_color:
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

    BLUE = "BDD7EE"
    YELLOW = "FFFF99"

    sub_labels = [
        (1,  "S.NO",                                    BLUE),
        (2,  "PDF NAME",                                BLUE),
        (3,  "Invoice No.\n(from Shipping\nBill)",      BLUE),
        (4,  "Port Code\n(from Shipping\nBill)",        BLUE),
        (5,  "Shipping Bill\nNumber\n(from Shipping\nBill)", BLUE),
        (6,  "Shipping Bill\nDate\n(from Shipping\nBill)",   BLUE),
        (7,  "Invoice Date\n(from Shipping\nBill)",     BLUE),
        (8,  "IGST\nVALUE",                             BLUE),
        (9,  "IGST TAX\nAMOUNT",                        BLUE),
        (10, "FOB VALUE",                               BLUE),
        (11, "FREIGHT",                                 BLUE),
        (12, "INSURANCE",                               BLUE),
        (13, "TAXABLE\nVALUE\n(FOB+FREIGHT\n+INSURANCE)", BLUE),
        (14, "LUT",                                     BLUE),
    ]

    for col, label, fill in sub_labels:
        c = ws.cell(row=1, column=col, value=label)
        style_cell(c, bold=True, font_color="000000", fill_color=fill, wrap=True)

    ws.row_dimensions[1].height = 65

    col_widths = {
        1: 6, 2: 35, 3: 18, 4: 11, 5: 14,
        6: 13, 7: 13, 8: 14, 9: 14, 10: 13,
        11: 11, 12: 11, 13: 16, 14: 8,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    for r_idx, row in enumerate(df.values.tolist()):
        excel_row = r_idx + 2

        s_no      = row[0]
        pdf_name  = row[1]
        inv_no    = row[2]
        port_code = row[3]
        sb_no     = row[4]
        sb_date   = row[5]
        inv_date  = row[6]
        tax_val   = row[7]
        igst_amt  = row[8]
        fob       = row[9]
        freight   = row[10]
        insurance = row[11]
        lut       = row[13]
        is_total  = (inv_no == "TOTAL")

        def val(v):
            if v == "" or v is None:
                return None
            return v

        values = [
            val(s_no), val(pdf_name), val(inv_no), val(port_code),
            val(sb_no), val(sb_date), val(inv_date), val(tax_val),
            val(igst_amt), val(fob), val(freight), val(insurance),
            None,
            val(lut),
        ]

        for col_idx, v in enumerate(values, start=1):
            c = ws.cell(row=excel_row, column=col_idx, value=v)
            c.border = bdr
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(bold=is_total, size=9)

        ws.row_dimensions[excel_row].height = 15

    last_data_row = 1 + len(df)
    summary_start = last_data_row + 5

    summary_headers = ["S.No.", "INV No.", "INV Amt.", "Currency", "Ex. Rate"]

    for col_idx, label in enumerate(summary_headers, start=1):
        c = ws.cell(row=summary_start, column=col_idx, value=label)
        style_cell(c, bold=True, font_color="000000", fill_color=YELLOW, wrap=False, size=9)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            col_widths.get(col_idx, 12), 14
        )

    ws.row_dimensions[summary_start].height = 18

    for s_idx, srow in enumerate(summary_rows, start=1):
        excel_row = summary_start + s_idx
        row_vals = [
            s_idx,
            srow.get("F_INV_NO"),
            srow.get("F_INV_AMT"),
            srow.get("F_CURRENCY"),
            srow.get("EXCHANGE_RATE"),
        ]
        for col_idx, v in enumerate(row_vals, start=1):
            c = ws.cell(row=excel_row, column=col_idx, value=v)
            c.border = bdr
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=9)
        ws.row_dimensions[excel_row].height = 15

    wb.save(output)
    output.seek(0)
    return output


uploaded_files = st.file_uploader(
    "Upload PDF Files", type=["pdf"], accept_multiple_files=True
)

debug_mode = st.checkbox("🔍 Debug Mode (show raw page texts)")

if st.button("Extract Data"):
    if not uploaded_files:
        st.warning("Please upload PDF files.")
    else:
        results = []
        summary_rows = []
        duplicate_checker = set()
        progress_bar = st.progress(0)

        for index, uploaded_file in enumerate(uploaded_files):
            progress_bar.progress((index + 1) / len(uploaded_files))
            st.write(f"Processing: {uploaded_file.name}")

            full_text, page_texts = extract_pdf_text(uploaded_file)

            if debug_mode:
                st.markdown(f"**--- PAGE 1 RAW TEXT ({uploaded_file.name}) ---**")
                st.text(page_texts[0] if page_texts else "No page 1 text")
                st.markdown(f"**--- PAGE 2 RAW TEXT ({uploaded_file.name}) ---**")
                st.text(page_texts[1] if len(page_texts) > 1 else "No page 2 found")

            extracted = extract_data(full_text, page_texts)
            extracted["PDF Name"] = uploaded_file.name

            invoice_no = extracted["Invoice No. (from Shipping Bill)"]
            if invoice_no in duplicate_checker and invoice_no != "":
                continue
            duplicate_checker.add(invoice_no)

            summary_rows.append({
                "F_INV_NO":      extracted.get("F_INV_NO"),
                "F_INV_AMT":     extracted.get("F_INV_AMT"),
                "F_CURRENCY":    extracted.get("F_CURRENCY"),
                "EXCHANGE_RATE": extracted.get("EXCHANGE_RATE"),
            })

            results.append(extracted)

        columns = [
            "S NO.",
            "PDF Name",
            "Invoice No. (from Shipping Bill)",
            "Port Code (from Shipping Bill)",
            "Shipping Bill No. (from Shipping Bill)",
            "Shipping Bill Date (from Shipping Bill)",
            "Invoice Date (from Shipping Bill)",
            "IGST VALUE",
            "IGST TAX AMOUNT",
            "FOB",
            "FREIGHT",
            "INSURANCE",
            "TAXABLE VALUE (FOB+FREIGHT+INSURANCE)",
            "LUT"
        ]

        df = pd.DataFrame(results)

        df["SORT_DATE"] = pd.to_datetime(
            df["Invoice Date (from Shipping Bill)"],
            format="%d-%b-%y",
            errors="coerce"
        )

        df["SORT_INV"] = (
            df["Invoice No. (from Shipping Bill)"]
            .str.extract(r'(\d+)$')[0]
            .astype(float)
        )

        df = (
            df.sort_values(
                by=["SORT_DATE", "SORT_INV"],
                ascending=[True, True]
            )
            .drop(columns=["SORT_DATE", "SORT_INV"])
        )

        df.insert(0, "S NO.", range(1, len(df) + 1))
        df = df[columns]

        def safe_sum(col):
            vals = pd.to_numeric(df[col], errors='coerce').fillna(0)
            s = round(vals.sum(), 2)
            return s if s != 0 else ""

        total_taxable   = safe_sum("IGST VALUE")
        total_igst      = safe_sum("IGST TAX AMOUNT")
        total_fob       = safe_sum("FOB")
        total_freight   = safe_sum("FREIGHT")
        total_insurance = safe_sum("INSURANCE")

        empty_rows = pd.DataFrame([[None] * len(columns)] * 4, columns=columns)

        total_row = pd.DataFrame([[
            None, None, "TOTAL", None, None, None, None,
            total_taxable, total_igst,
            total_fob, total_freight, total_insurance,
            None, None
        ]], columns=columns)

        df = pd.concat([df, empty_rows, total_row], ignore_index=True)

        invoice_order = {
            inv: idx
            for idx, inv in enumerate(df["Invoice No. (from Shipping Bill)"])
        }

        summary_rows = sorted(
            summary_rows,
            key=lambda x: invoice_order.get(x.get("F_INV_NO"), 999999)
        )

        output = build_excel(df, summary_rows)

        st.success("Extraction Completed!")
        st.dataframe(df[df["S NO."].notna()])

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="📥 Download Excel File",
            data=output,
            file_name=f"Shipping_Bill_Data_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
