import streamlit as st
import pdfplumber
import pandas as pd
import re
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# PAGE SETTINGS
# ==========================================

st.set_page_config(page_title="PDF Extractor", layout="wide")
st.title("📄 GST Shipping Bill PDF Extractor")
st.write("Upload PDF files and extract invoice data into Excel.")

# ==========================================
# EXTRACT TEXT FROM PDF
# ==========================================

def extract_pdf_text(uploaded_file):
    full_text = ""
    try:
        with pdfplumber.open(uploaded_file) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
    except Exception as e:
        st.error(f"PDF Error: {e}")
    return full_text

# ==========================================
# EXTRACT REQUIRED DATA
# ==========================================

def extract_data(text):
    data = {
        "Invoice No. (from Shipping Bill)": None,
        "Port Code (from Shipping Bill)": None,
        "Shipping Bill No. (from Shipping Bill)": None,
        "Shipping Bill Date (from Shipping Bill)": None,
        "Invoice Date (from Shipping Bill)": None,
        "TAXABLE VALUE": None,
        "IGST TAX AMOUNT": None,
        "FOB": None,
        "FREIGHT": None,
        "INSURANCE": None,
        "TAXABLE VALUE (FOB+FREIGHT+INSURANCE)": None,
        "LUT": None
    }

    clean_text = re.sub(r'\s+', ' ', text)

    # --- Port Code, SB No, SB Date ---
    sb_match = re.search(
        r'INDIAN CUSTOMS EDI SYSTEM\s+([A-Z0-9]+)\s+(\d+)\s+([0-9A-Z\-]+)',
        clean_text
    )
    if sb_match:
        data["Port Code (from Shipping Bill)"] = sb_match.group(1)
        data["Shipping Bill No. (from Shipping Bill)"] = sb_match.group(2)
        data["Shipping Bill Date (from Shipping Bill)"] = sb_match.group(3)

    # --- Invoice No. & Date from PART-II A.REF ---
    inv_match = re.search(
        r'(JTIPL/\d{4}/\d+)\s+(\d{2}/\d{2}/\d{4})',
        clean_text
    )
    if inv_match:
        data["Invoice No. (from Shipping Bill)"] = inv_match.group(1)
        original_date = inv_match.group(2)
        try:
            converted_date = datetime.strptime(original_date, "%d/%m/%Y").strftime("%d-%b-%y").upper()
            data["Invoice Date (from Shipping Bill)"] = converted_date
        except ValueError:
            data["Invoice Date (from Shipping Bill)"] = original_date

    # --- FOB Value (LM pattern = assessed FOB in INR) ---
    fob_match = re.search(r'LM\s+([0-9.]+)', clean_text)
    fob_value = ""
    if fob_match:
        fob_value = round(float(fob_match.group(1)), 2)
        data["FOB"] = fob_value
        data["TAXABLE VALUE"] = fob_value

    # --- Freight ---
    freight_match = re.search(
        r'5\.COM 2\. IGST AMT\s+[0-9]+(?:\.[0-9]+)?\s+([0-9]+(?:\.[0-9]+)?)',
        clean_text
    )
    freight_value = ""
    if freight_match:
        v = float(freight_match.group(1))
        if v > 0:
            freight_value = round(v, 2)
            data["FREIGHT"] = freight_value

    # --- Insurance ---
    insurance_match = re.search(
        r'5\.COM 2\. IGST AMT\s+[0-9]+(?:\.[0-9]+)?\s+[0-9]+(?:\.[0-9]+)?\s+([0-9]+(?:\.[0-9]+)?)',
        clean_text
    )
    insurance_value = ""
    if insurance_match:
        v = float(insurance_match.group(1))
        if v > 0:
            insurance_value = round(v, 2)
            data["INSURANCE"] = insurance_value

    # # --- TAXABLE VALUE (FOB+FREIGHT+INSURANCE) ---
    # fob_n = fob_value if isinstance(fob_value, float) else 0.0
    # freight_n = freight_value if isinstance(freight_value, float) else 0.0
    # insurance_n = insurance_value if isinstance(insurance_value, float) else 0.0
    # total = round(fob_n + freight_n + insurance_n, 2)
    # data["TAXABLE VALUE (FOB+FREIGHT+INSURANCE)"] = total if total > 0 else None
    
    # Leave TAXABLE VALUE (FOB+FREIGHT+INSURANCE) blank
    data["TAXABLE VALUE (FOB+FREIGHT+INSURANCE)"] = None
    
    # --- LUT detection: if 4.IGST VALUE is followed by a number -> LUT=N, else LUT=Y ---
    igst_val_present = re.search(r'4\.IGST VALUE\s+(\d[\d.]*)', clean_text)
    is_lut = not bool(igst_val_present)
    data["LUT"] = "Y" if is_lut else "N"

    # --- IGST Amount: only for non-LUT shipments ---
    if not is_lut:
        igst_match = re.search(r'3\.CESS AMT\s+([0-9.]+)\s+([0-9.]+)', clean_text)
        if igst_match:
            igst_amt = round(float(igst_match.group(2)), 2)
            data["IGST TAX AMOUNT"] = igst_amt if igst_amt > 0 else None

    return data

# ==========================================
# BUILD EXCEL — SUBHEADINGS ONLY, NO GROUP HEADERS
# ==========================================

def build_excel(df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_cell(cell, bold=True, font_color="000000", fill_color=None,
                   wrap=True, size=9, italic=False, h_align="center"):
        cell.font      = Font(bold=bold, color=font_color, size=size, italic=italic)
        cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
        cell.border    = bdr
        if fill_color:
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

    # ---- COLOURS ----
    BLUE   = "BDD7EE"
    YELLOW = "FFFF99"
    GREEN  = "C6EFCE"
    ORANGE = "FCE4D6"
    PURPLE = "E2EFDA"
    GRAY   = "D9D9D9"

    # ---- COLUMN LAYOUT (20 columns) ----
    # 1:S.NO | 2-6:SB details | 7-8:Invoice/E-Invoice | 9-11:FOB/FREIGHT/INSURANCE
    # 12:Shipping Bill taxable | 13:blank | 14-15:Differences | 16-17:Shut Out Diff
    # 18-20:GSTR-1 report

    sub_labels = [
    (1,  "S.NO", BLUE),
    (2,  "PDF NAME", BLUE),
    (3,  "Invoice No.\n(from Shipping\nBill)", BLUE),
    (4,  "Port Code\n(from Shipping\nBill)", BLUE),
    (5,  "Shipping Bill\nNumber\n(from Shipping\nBill)", BLUE),
    (6,  "Shipping Bill\nDate\n(from Shipping\nBill)", BLUE),
    (7,  "Invoice Date\n(from Shipping\nBill)", BLUE),
    (8,  "TAXABLE\nVALUE", BLUE),
    (9,  "IGST TAX\nAMOUNT", BLUE),
    (10, "FOB VALUE", BLUE),
    (11, "FREIGHT", BLUE),
    (12, "INSURANCE", BLUE),
    (13, "TAXABLE\nVALUE\n(FOB+FREIGHT\n+INSURANCE)", BLUE),
    (14, "LUT", BLUE),
]

    # Row 1 = subheadings only
    for col, label, fill in sub_labels:
        c = ws.cell(row=1, column=col, value=label)
        style_cell(c, bold=True, font_color="FF0000", fill_color=fill, wrap=True)

    ws.row_dimensions[1].height = 65

    col_widths = {
    1: 6,
    2: 35,   # PDF Name
    3: 18,
    4: 11,
    5: 14,
    6: 13,
    7: 13,
    8: 14,
    9: 14,
    10: 13,
    11: 11,
    12: 11,
    13: 16,
    14: 8,
}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ---- DATA ROWS (start row 2) ----
    for r_idx, row in enumerate(df.values.tolist()):
        excel_row = r_idx + 2

        s_no      = row[0]
        pdf_name  = row[1]
        inv_no    = row[2]
        port_code = row[3]
        sb_no     = row[4]
        sb_date   = row[5]
        inv_date  = row[6]
        tax_val   = row[7]
        igst_amt  = row[8]
        fob       = row[9]
        freight   = row[10]
        insurance = row[11]
        lut       = row[13]
        is_total = (inv_no == "TOTAL")

        # Blank zeros — keep empty string as None
        def val(v):
            if v == "" or v is None:
                return None
            return v

        values = [
    val(s_no),
    val(pdf_name),
    val(inv_no),
    val(port_code),
    val(sb_no),
    val(sb_date),
    val(inv_date),
    val(tax_val),
    val(igst_amt),
    val(fob),
    val(freight),
    val(insurance),
    None,  # val(tax_val_fob),  # Leave TAXABLE VALUE (FOB+FREIGHT+INSURANCE) blank
    val(lut),
]

        for col_idx, v in enumerate(values, start=1):
            c = ws.cell(row=excel_row, column=col_idx, value=v)
            c.border    = bdr
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font      = Font(bold=is_total, size=9)

        ws.row_dimensions[excel_row].height = 15

    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# FILE UPLOAD & EXTRACT
# ==========================================

uploaded_files = st.file_uploader(
    "Upload PDF Files", type=["pdf"], accept_multiple_files=True
)

if st.button("Extract Data"):
    if not uploaded_files:
        st.warning("Please upload PDF files.")
    else:
        results = []
        duplicate_checker = set()
        progress_bar = st.progress(0)

        for index, uploaded_file in enumerate(uploaded_files):
            progress_bar.progress((index + 1) / len(uploaded_files))
            st.write(f"Processing: {uploaded_file.name}")
            text      = extract_pdf_text(uploaded_file)
            extracted = extract_data(text)
            extracted["PDF Name"] = uploaded_file.name
            invoice_no = extracted["Invoice No. (from Shipping Bill)"]
            if invoice_no in duplicate_checker and invoice_no != "":
                continue
            duplicate_checker.add(invoice_no)
            results.append(extracted)

        # ---- DATAFRAME ----
        columns = [
            "S NO.",
            "PDF Name",
            "Invoice No. (from Shipping Bill)",
            "Port Code (from Shipping Bill)",
            "Shipping Bill No. (from Shipping Bill)",
            "Shipping Bill Date (from Shipping Bill)",
            "Invoice Date (from Shipping Bill)",
            "TAXABLE VALUE",
            "IGST TAX AMOUNT",
            "FOB",
            "FREIGHT",
            "INSURANCE",
            "TAXABLE VALUE (FOB+FREIGHT+INSURANCE)",
            "LUT"
        ]

        df = pd.DataFrame(results)

        df["SORT_DATE"] = pd.to_datetime(
            df["Invoice Date (from Shipping Bill)"],
            format="%d-%b-%y", errors="coerce"
        )
        df = df.sort_values(by="SORT_DATE", ascending=True).drop(columns=["SORT_DATE"])
        df.insert(0, "S NO.", range(1, len(df) + 1))
        df = df[columns]

        # ---- TOTALS ----
        def safe_sum(col):
            vals = pd.to_numeric(df[col], errors='coerce').fillna(0)
            s = round(vals.sum(), 2)
            return s if s != 0 else ""

        total_taxable   = safe_sum("TAXABLE VALUE")
        total_igst      = safe_sum("IGST TAX AMOUNT")
        total_fob       = safe_sum("FOB")
        total_freight   = safe_sum("FREIGHT")
        total_insurance = safe_sum("INSURANCE")
        # total_tax_fob   = safe_sum("TAXABLE VALUE (FOB+FREIGHT+INSURANCE)")

        empty_rows = pd.DataFrame([[None] * len(columns)] * 4, columns=columns)

        total_row = pd.DataFrame([[
            None,None, "TOTAL", None, None, None, None,
            total_taxable, total_igst,
            total_fob, total_freight, total_insurance,
            # total_tax_fob,
            None,
            None
        ]], columns=columns)

        df = pd.concat([df, empty_rows, total_row], ignore_index=True)

        # ---- BUILD & DOWNLOAD ----
        output = build_excel(df)

        st.success("Extraction Completed!")
        st.dataframe(df[df["S NO."].notna()])

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="📥 Download Excel File",
            data=output,
            file_name=f"Shipping_Bill_Data_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )